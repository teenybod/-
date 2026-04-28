import os
import json
import io
import uuid
import requests
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, send_file, g

from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from models import (
    db, Process, FilterModel, Filter, UsageRecord,
    SterilizationRecord, Config, User, FilterReplacementRecord
)

app = Flask(__name__)

# 环境检测
is_vercel = os.environ.get('VERCEL') == '1' or os.environ.get('VERCEL_ENV') is not None

# 数据库配置：优先使用环境变量 DATABASE_URL（PostgreSQL / MySQL / SQLite）
db_url = os.environ.get('DATABASE_URL')
if db_url:
    # 部分平台（如 Supabase/Heroku）使用 postgres:// 前缀，SQLAlchemy 需要 postgresql://
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
elif is_vercel:
    # Vercel 无外部数据库时的降级方案（数据不持久，仅供演示）
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////tmp/filter_mgmt.db'
else:
    # 本地开发 / 宝塔 Linux 部署：使用项目目录下的绝对路径，避免工作目录变化导致找不到数据库
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'filter_mgmt.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JSON_AS_ASCII'] = False
app.secret_key = os.environ.get('SECRET_KEY', 'filter_mgmt_secret_key_2024')

db.init_app(app)

# ==================== 定时推送 ====================
scheduler = BackgroundScheduler()


def _get_usage_alert_status(f):
    """判断使用管理物料的预警状态"""
    alerts = {'sterilization': False, 'expiry': False, 'messages': [], 'remaining': None, 'days_left': None}

    cfg = Config.query.first()
    alert_days = cfg.alert_days if cfg else 7
    alert_sterilization_remaining = cfg.alert_sterilization_remaining if cfg else 0

    # 灭菌次数预警：基于物料型号的 max_sterilization_count
    max_ster = None
    if f.model and f.model.max_sterilization_count is not None:
        max_ster = f.model.max_sterilization_count
    if max_ster and max_ster > 0:
        remaining = max_ster - f.current_sterilization_count
        alerts['remaining'] = remaining
        if remaining <= alert_sterilization_remaining:
            alerts['sterilization'] = True
            alerts['messages'].append(f'可灭菌次数仅剩 {remaining} 次')

    # 有效期预警：基于 production_date + 物料型号 max_days
    if f.model and f.model.max_days and f.model.max_days > 0 and f.production_date:
        expire_date = f.production_date + timedelta(days=f.model.max_days)
        days_left = (expire_date - datetime.now().date()).days
        alerts['days_left'] = days_left
        if days_left <= alert_days:
            alerts['expiry'] = True
            alerts['messages'].append(f'距有效期仅剩 {days_left} 天')

    return alerts


def _should_push_usage_alert(last_alert_at):
    """判断当天是否需要再次推送（08:00 和 20:00 各一次）"""
    if not last_alert_at:
        return True
    now = datetime.now()
    if last_alert_at.date() != now.date():
        return True
    if abs(now.hour - last_alert_at.hour) >= 6:
        return True
    return False


def auto_push_usage_alerts():
    """定时任务：推送使用管理物料的灭菌次数和有效期预警"""
    with app.app_context():
        cfg = Config.query.first()
        if not cfg or not cfg.feishu_webhook:
            return

        filters = Filter.query.filter_by(is_usage_record=True).all()
        alert_items = []

        for f in filters:
            status = _get_usage_alert_status(f)
            if not status['messages']:
                # 状态已恢复正常，重置推送记录
                if f.last_sterilization_alert_at or f.last_expiry_alert_at:
                    f.last_sterilization_alert_at = None
                    f.last_expiry_alert_at = None
                    db.session.commit()
                continue

            # 灭菌次数预警推送
            if status['sterilization'] and _should_push_usage_alert(f.last_sterilization_alert_at):
                alert_items.append({
                    'code': f.code,
                    'model_name': f.model.name if f.model else '',
                    'use_location': f.use_location or '',
                    'operator': f.operator or '',
                    'receivers': f.receivers or '',
                    'message': f"该物料（{f.code} {f.model.name if f.model else ''} {f.use_location or ''}）可灭菌次数仅剩1次，请及时更换"
                })
                f.last_sterilization_alert_at = datetime.now()
                db.session.commit()

            # 有效期预警推送
            if status['expiry'] and _should_push_usage_alert(f.last_expiry_alert_at):
                alert_items.append({
                    'code': f.code,
                    'model_name': f.model.name if f.model else '',
                    'use_location': f.use_location or '',
                    'operator': f.operator or '',
                    'receivers': f.receivers or '',
                    'message': f"该物料（{f.code} {f.model.name if f.model else ''} {f.use_location or ''}）距有效期仅剩{status['days_left']}天，请及时更换"
                })
                f.last_expiry_alert_at = datetime.now()
                db.session.commit()

        if not alert_items:
            return

        elements = []
        for alert in alert_items:
            elements.append({
                "tag": "div",
                "text": {"tag": "lark_md", "content": f"• **{alert['message']}**"}
            })
            if alert['operator'] or alert['receivers']:
                info = []
                if alert['operator']:
                    info.append(f"输入人：{alert['operator']}")
                if alert['receivers']:
                    info.append(f"接收人：{alert['receivers']}")
                elements.append({
                    "tag": "div",
                    "text": {"tag": "lark_md", "content": f"  {' | '.join(info)}"}
                })

        card = {
            "config": {"wide_screen_mode": True},
            "header": {
                "title": {"tag": "plain_text", "content": f"物料使用预警 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"},
                "template": "yellow"
            },
            "elements": elements
        }

        payload = {"msg_type": "interactive", "card": card}
        try:
            requests.post(cfg.feishu_webhook, json=payload, timeout=15)
        except Exception:
            pass


# Serverless 环境（如 Vercel）不支持常驻后台任务，只在传统服务器上启动定时器
if not is_vercel:
    scheduler.add_job(auto_push_usage_alerts, 'cron', hour='8,20', minute='0')
    scheduler.start()

# ==================== 初始化数据 ====================

def init_data():
    with app.app_context():
        db.create_all()
        # 默认工序
        if not Process.query.first():
            default_processes = ['洗瓶', '消毒', '配料', '灌轧', '冻干']
            for name in default_processes:
                db.session.add(Process(name=name))
            db.session.commit()
        # 默认配置
        if not Config.query.first():
            db.session.add(Config())
            db.session.commit()
        # 默认管理员
        if not User.query.filter_by(username='admin').first():
            u = User(username='admin', real_name='管理员', role='admin')
            u.set_password('1234')
            db.session.add(u)
            db.session.commit()


init_data()


# ==================== 飞书扫码登录 ====================
# 内存缓存：state -> {'status': 'pending'|'success'|'error', 'user': user_dict, 'msg': ''}
_pending_feishu_logins = {}


def _get_feishu_app_access_token(app_id, app_secret):
    """获取飞书 app_access_token（内部应用）"""
    url = 'https://open.feishu.cn/open-apis/auth/v3/app_access_token/internal'
    resp = requests.post(url, json={'app_id': app_id, 'app_secret': app_secret}, timeout=15)
    data = resp.json()
    if data.get('code') == 0:
        return data.get('app_access_token')
    return None


def _get_feishu_tenant_access_token(app_id, app_secret):
    """获取飞书 tenant_access_token（用于 Bitable API）"""
    url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
    resp = requests.post(url, json={'app_id': app_id, 'app_secret': app_secret}, timeout=15)
    data = resp.json()
    if data.get('code') == 0:
        return data.get('tenant_access_token')
    return None


# ==================== 飞书多维表格同步 ====================

def _feishu_bitable_headers(token):
    return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}


def _get_feishu_bitable_tables(app_token, tenant_token):
    """获取多维表格下的所有表格，返回 table_id 列表"""
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables'
    resp = requests.get(url, headers=_feishu_bitable_headers(tenant_token), timeout=15)
    data = resp.json()
    if data.get('code') == 0:
        tables = data.get('data', {}).get('items', [])
        return [(t.get('table_id'), t.get('name')) for t in tables]
    return []


def _get_feishu_bitable_records(app_token, table_id, tenant_token, page_size=500):
    """分页获取多维表格所有记录"""
    records = []
    page_token = None
    while True:
        url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records?page_size={page_size}'
        if page_token:
            url += f'&page_token={page_token}'
        resp = requests.get(url, headers=_feishu_bitable_headers(tenant_token), timeout=30)
        data = resp.json()
        if data.get('code') != 0:
            break
        items = data.get('data', {}).get('items', [])
        records.extend(items)
        if not data.get('data', {}).get('has_more'):
            break
        page_token = data.get('data', {}).get('page_token')
    return records


def _create_feishu_bitable_record(app_token, table_id, tenant_token, fields):
    """在飞书多维表格创建记录"""
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records'
    resp = requests.post(url, headers=_feishu_bitable_headers(tenant_token), json={'fields': fields}, timeout=15)
    return resp.json()


def _update_feishu_bitable_record(app_token, table_id, record_id, tenant_token, fields):
    """更新飞书多维表格记录"""
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/{record_id}'
    resp = requests.put(url, headers=_feishu_bitable_headers(tenant_token), json={'fields': fields}, timeout=15)
    return resp.json()


def _delete_feishu_bitable_record(app_token, table_id, record_id, tenant_token):
    """删除飞书多维表格记录"""
    url = f'https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/{record_id}'
    resp = requests.delete(url, headers=_feishu_bitable_headers(tenant_token), timeout=15)
    return resp.json()


def _parse_feishu_record(fields):
    """解析飞书记录字段为系统字典"""
    # 飞书字段值可能是文本或数字，处理嵌套结构
    def _val(v):
        if isinstance(v, list) and len(v) > 0:
            # 多行文本/链接/人员等字段取第一个元素的 text
            first = v[0]
            if isinstance(first, dict):
                return first.get('text', first.get('label', str(first)))
            return str(first)
        if isinstance(v, dict):
            return v.get('text', v.get('value', str(v)))
        return v

    code = str(_val(fields.get('物料代码', ''))).strip()
    name = str(_val(fields.get('物料名称', ''))).strip()
    spec = str(_val(fields.get('型号/规格', ''))).strip()
    unit = str(_val(fields.get('单位', ''))).strip()
    supplier = str(_val(fields.get('供应商名称', ''))).strip()

    # 可灭菌次数
    max_ster = _val(fields.get('可灭菌次数', None))
    try:
        max_sterilization_count = int(float(max_ster)) if max_ster is not None and str(max_ster).strip() != '' else None
    except (ValueError, TypeError):
        max_sterilization_count = None

    # 有效期(年) → 天数
    max_years = _val(fields.get('有效期(年)', None))
    try:
        years = float(max_years) if max_years is not None and str(max_years).strip() != '' else 0
        max_days = int(years * 365) if years > 0 else None
    except (ValueError, TypeError):
        max_days = None

    return {
        'code': code,
        'name': name,
        'spec': spec,
        'unit': unit,
        'supplier': supplier,
        'max_sterilization_count': max_sterilization_count,
        'max_days': max_days,
    }


def _build_feishu_fields(filter_obj, model_obj):
    """将系统对象构建为飞书多维表格字段字典"""
    fields = {
        '物料代码': filter_obj.code,
        '物料名称': model_obj.name if model_obj else '',
        '型号/规格': model_obj.spec if model_obj else '',
        '单位': model_obj.unit or '',
        '供应商名称': model_obj.supplier or '',
    }
    if model_obj and model_obj.max_sterilization_count is not None:
        fields['可灭菌次数'] = model_obj.max_sterilization_count
    if model_obj and model_obj.max_days is not None and model_obj.max_days > 0:
        fields['有效期(年)'] = round(model_obj.max_days / 365, 2)
    return fields


def sync_from_feishu_bitable():
    """从飞书多维表格同步数据到系统（定时任务或手动调用）"""
    cfg = Config.query.first()
    if not cfg or not cfg.feishu_app_id or not cfg.feishu_app_secret:
        return {'success': False, 'message': '未配置飞书应用凭证'}
    if not cfg.feishu_bitable_app_token:
        return {'success': False, 'message': '未配置飞书多维表格 App Token'}

    tenant_token = _get_feishu_tenant_access_token(cfg.feishu_app_id, cfg.feishu_app_secret)
    if not tenant_token:
        return {'success': False, 'message': '无法获取飞书 tenant_access_token'}

    app_token = cfg.feishu_bitable_app_token
    table_id = cfg.feishu_bitable_table_id

    # 如果没有配置 table_id，尝试自动获取第一个表格
    if not table_id:
        tables = _get_feishu_bitable_tables(app_token, tenant_token)
        if tables:
            table_id = tables[0][0]
            cfg.feishu_bitable_table_id = table_id
            db.session.commit()
        else:
            return {'success': False, 'message': '无法获取多维表格的表格列表'}

    records = _get_feishu_bitable_records(app_token, table_id, tenant_token)
    if not records:
        return {'success': True, 'message': '飞书表格暂无数据', 'created': 0, 'updated': 0, 'deleted': 0}

    created_count = 0
    updated_count = 0

    # 收集飞书中所有的物料代码和 record_id
    feishu_codes = {}
    feishu_record_ids = set()

    for rec in records:
        record_id = rec.get('record_id')
        fields = rec.get('fields', {})
        parsed = _parse_feishu_record(fields)
        code = parsed['code']
        if not code:
            continue
        feishu_codes[code] = {
            'record_id': record_id,
            'parsed': parsed,
            'updated_time': rec.get('last_modified_time', ''),
        }
        feishu_record_ids.add(record_id)

    # 遍历飞书记录，同步到系统
    for code, info in feishu_codes.items():
        parsed = info['parsed']
        record_id = info['record_id']
        feishu_updated_str = info['updated_time']

        # 查找或创建 FilterModel
        model = FilterModel.query.filter_by(name=parsed['name']).first() if parsed['name'] else None
        if not model and parsed['name']:
            model = FilterModel(name=parsed['name'])
            db.session.add(model)
            db.session.flush()

        if model:
            # 冲突判断：比较飞书和系统的更新时间
            feishu_dt = None
            try:
                feishu_dt = datetime.fromisoformat(feishu_updated_str.replace('Z', '+00:00'))
            except Exception:
                pass

            sys_dt = model.updated_at
            if feishu_dt and sys_dt and sys_dt.tzinfo is None:
                # 本地时间视为 UTC+8 进行比较（简化处理）
                pass

            # 如果飞书较新或系统无记录，则更新型号数据
            should_update_model = True
            if feishu_dt and sys_dt and sys_dt > feishu_dt.replace(tzinfo=None):
                should_update_model = False

            if should_update_model:
                model.spec = parsed['spec'] or model.spec
                model.unit = parsed['unit'] or model.unit
                model.supplier = parsed['supplier'] or model.supplier
                if parsed['max_sterilization_count'] is not None:
                    model.max_sterilization_count = parsed['max_sterilization_count']
                if parsed['max_days'] is not None:
                    model.max_days = parsed['max_days']
                model.feishu_record_id = record_id
                model.updated_at = datetime.now()

        # 查找或创建 Filter
        filter_obj = Filter.query.filter_by(code=code).first()
        if not filter_obj:
            filter_obj = Filter(
                code=code,
                model_id=model.id if model else None,
                start_date=datetime.now().date(),
                status='normal',
                current_usage_count=0,
                current_sterilization_count=0,
            )
            db.session.add(filter_obj)
            db.session.flush()
            created_count += 1
        else:
            # 更新关联的型号
            if model and filter_obj.model_id != model.id:
                filter_obj.model_id = model.id
            filter_obj.feishu_record_id = record_id
            filter_obj.updated_at = datetime.now()
            updated_count += 1

    db.session.commit()

    # 双向同步：将系统较新的数据推回飞书
    # 只推送那些在系统中被修改且比飞书新的记录
    push_count = 0
    local_filters = Filter.query.filter(Filter.feishu_record_id.isnot(None)).all()
    for f in local_filters:
        if not f.model:
            continue
        feishu_info = None
        for code, info in feishu_codes.items():
            if info['record_id'] == f.feishu_record_id:
                feishu_info = info
                break
        if not feishu_info:
            continue
        feishu_dt = None
        try:
            feishu_dt = datetime.fromisoformat(feishu_info['updated_time'].replace('Z', '+00:00'))
        except Exception:
            pass
        sys_dt = f.model.updated_at or f.updated_at
        if sys_dt and feishu_dt and sys_dt > feishu_dt.replace(tzinfo=None):
            fields = _build_feishu_fields(f, f.model)
            result = _update_feishu_bitable_record(app_token, table_id, f.feishu_record_id, tenant_token, fields)
            if result.get('code') == 0:
                push_count += 1

    return {
        'success': True,
        'message': f'同步完成：新增 {created_count} 条，更新 {updated_count} 条，推送 {push_count} 条',
        'created': created_count,
        'updated': updated_count,
        'pushed': push_count,
    }


def auto_sync_feishu_bitable():
    """定时任务：自动同步飞书多维表格"""
    with app.app_context():
        try:
            result = sync_from_feishu_bitable()
            print(f'[AutoSync] {result}')
        except Exception as e:
            print(f'[AutoSync] Error: {e}')



def _get_feishu_user_info(code, app_access_token):
    """用授权码换取飞书用户信息"""
    # 1. 获取 user_access_token
    token_url = 'https://open.feishu.cn/open-apis/authen/v1/oidc/access_token'
    resp = requests.post(
        token_url,
        headers={'Authorization': f'Bearer {app_access_token}', 'Content-Type': 'application/json'},
        json={'grant_type': 'authorization_code', 'code': code},
        timeout=15
    )
    token_data = resp.json()
    if token_data.get('code') != 0:
        return None, token_data.get('msg', '获取 access_token 失败')
    user_access_token = token_data.get('data', {}).get('access_token')
    if not user_access_token:
        return None, 'access_token 为空'

    # 2. 获取用户信息
    info_url = 'https://open.feishu.cn/open-apis/authen/v1/user_info'
    resp = requests.get(
        info_url,
        headers={'Authorization': f'Bearer {user_access_token}'},
        timeout=15
    )
    info_data = resp.json()
    if info_data.get('code') != 0:
        return None, info_data.get('msg', '获取用户信息失败')
    return info_data.get('data'), None


@app.route('/api/feishu/auth_url')
def feishu_auth_url():
    """生成飞书授权 URL，供前端生成二维码"""
    cfg = Config.query.first()
    if not cfg or not cfg.feishu_app_id:
        return jsonify({'success': False, 'message': '未配置飞书应用 ID'}), 400

    state = str(uuid.uuid4()).replace('-', '')
    redirect_uri = request.url_root.rstrip('/') + '/feishu/callback'
    auth_url = (
        f"https://open.feishu.cn/open-apis/authen/v1/index"
        f"?app_id={cfg.feishu_app_id}"
        f"&redirect_uri={requests.utils.quote(redirect_uri, safe='')}"
        f"&state={state}"
    )

    _pending_feishu_logins[state] = {
        'status': 'pending',
        'created_at': datetime.now(),
        'user': None,
        'msg': ''
    }
    return jsonify({
        'success': True,
        'auth_url': auth_url,
        'state': state
    })


@app.route('/feishu/callback')
def feishu_callback():
    """飞书授权回调（手机端扫码后跳转至此）"""
    code = request.args.get('code', '')
    state = request.args.get('state', '')
    if not code:
        return '<h3>登录失败：未获取到授权码</h3>', 400

    cfg = Config.query.first()
    if not cfg or not cfg.feishu_app_id or not cfg.feishu_app_secret:
        return '<h3>登录失败：服务器未配置飞书应用</h3>', 500

    # 获取 app_access_token
    app_access_token = _get_feishu_app_access_token(cfg.feishu_app_id, cfg.feishu_app_secret)
    if not app_access_token:
        return '<h3>登录失败：无法获取应用凭证</h3>', 500

    # 获取用户信息
    user_info, err = _get_feishu_user_info(code, app_access_token)
    if err:
        return f'<h3>登录失败：{err}</h3>', 400

    open_id = user_info.get('open_id')
    union_id = user_info.get('union_id')
    name = user_info.get('name', '')
    # 优先使用 open_id 查找用户
    user = User.query.filter_by(feishu_open_id=open_id).first()
    if not user and union_id:
        user = User.query.filter_by(feishu_union_id=union_id).first()

    if user:
        if not user.is_active:
            return '<h3>登录失败：账号已被禁用</h3>', 403
        user.feishu_open_id = open_id
        user.feishu_union_id = union_id
        db.session.commit()
    else:
        # 自动创建用户（operator 角色）
        username = f"feishu_{open_id[-8:]}" if open_id else f"feishu_{uuid.uuid4().hex[:8]}"
        # 避免用户名冲突
        while User.query.filter_by(username=username).first():
            username = f"feishu_{uuid.uuid4().hex[:8]}"
        user = User(
            username=username,
            real_name=name or username,
            role='operator',
            feishu_open_id=open_id,
            feishu_union_id=union_id
        )
        user.set_password(uuid.uuid4().hex)
        db.session.add(user)
        db.session.commit()

    # 记录登录状态（供 PC 端轮询）
    login_payload = {
        'status': 'success',
        'user_id': user.id,
        'username': user.username,
        'real_name': user.real_name,
        'role': user.role
    }
    if state and state in _pending_feishu_logins:
        _pending_feishu_logins[state].update({
            'status': 'success',
            'user': login_payload
        })
    else:
        # 如果没有 state 或 state 过期，直接设置一个一次性 token
        token = str(uuid.uuid4()).replace('-', '')
        _pending_feishu_logins[token] = {
            'status': 'success',
            'user': login_payload,
            'created_at': datetime.now()
        }
        state = token

    # 返回一个简洁的 HTML，提示用户已授权，可以关闭页面
    return f"""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>飞书登录成功</title>
        <style>
            body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                   text-align: center; padding: 40px 20px; background: #f5f5f5; }}
            .card {{ background: white; border-radius: 12px; padding: 30px; max-width: 360px; margin: 0 auto;
                    box-shadow: 0 2px 12px rgba(0,0,0,0.08); }}
            .icon {{ font-size: 48px; color: #52c41a; margin-bottom: 16px; }}
            h2 {{ margin: 0 0 8px; color: #333; font-size: 20px; }}
            p {{ color: #666; margin: 0 0 20px; font-size: 14px; }}
            .btn {{ display: inline-block; background: #3370ff; color: white; padding: 10px 24px;
                    border-radius: 6px; text-decoration: none; font-size: 14px; }}
        </style>
    </head>
    <body>
        <div class="card">
            <div class="icon">✅</div>
            <h2>授权成功</h2>
            <p>您已成功登录，请返回电脑端继续操作。</p>
            <a href="/?feishu_state={state}" class="btn">进入系统</a>
        </div>
    </body>
    </html>
    """


@app.route('/api/feishu/login_status')
def feishu_login_status():
    """PC 端轮询登录状态"""
    state = request.args.get('state', '')
    if not state or state not in _pending_feishu_logins:
        return jsonify({'success': False, 'message': '无效的登录状态'}), 400

    record = _pending_feishu_logins[state]
    if record['status'] == 'pending':
        return jsonify({'success': True, 'status': 'pending'})

    if record['status'] == 'success':
        user_data = record['user']
        session['user_id'] = user_data['user_id']
        session['username'] = user_data['username']
        session['role'] = user_data['role']
        # 清理
        del _pending_feishu_logins[state]
        return jsonify({'success': True, 'status': 'success', 'data': user_data})

    return jsonify({'success': False, 'status': 'error', 'message': record.get('msg', '登录失败')})


# ==================== 登录与权限 ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json:
                return jsonify({'success': False, 'message': '请先登录'}), 401
            return render_template('login.html')
        g.user = User.query.get(session['user_id'])
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json:
                return jsonify({'success': False, 'message': '请先登录'}), 401
            return render_template('login.html')
        g.user = User.query.get(session['user_id'])
        if not g.user or g.user.role != 'admin':
            if request.is_json:
                return jsonify({'success': False, 'message': '权限不足，需要管理员权限'}), 403
            return render_template('login.html')
        return f(*args, **kwargs)
    return decorated_function


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    user = User.query.filter_by(username=username, is_active=True).first()
    if not user or not user.check_password(password):
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 400
    session['user_id'] = user.id
    session['username'] = user.username
    session['role'] = user.role
    return jsonify({'success': True, 'data': {'username': user.username, 'role': user.role, 'real_name': user.real_name}})


@app.route('/logout')
def logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/me')
@login_required
def get_me():
    return jsonify(g.user.to_dict())


# ==================== 页面路由 ====================

@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/processes')
@login_required
def processes_page():
    return render_template('processes.html')


@app.route('/models')
@login_required
def models_page():
    return render_template('models.html')


@app.route('/filters')
@login_required
def filters_page():
    return render_template('filters.html')


@app.route('/logs')
@login_required
def logs_page():
    return render_template('logs.html')


@app.route('/settings')
@admin_required
def settings_page():
    return render_template('settings.html')


@app.route('/reports')
@login_required
def reports_page():
    return render_template('reports.html')


# ==================== API: 用户管理 ====================

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    items = User.query.all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    if not username:
        return jsonify({'success': False, 'message': '用户名不能为空'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    u = User(
        username=username,
        real_name=data.get('real_name', ''),
        role=data.get('role', 'operator')
    )
    u.set_password(data.get('password', '123456'))
    db.session.add(u)
    db.session.commit()
    return jsonify({'success': True, 'data': u.to_dict()})


@app.route('/api/users/<int:uid>', methods=['PUT'])
@admin_required
def update_user(uid):
    u = User.query.get(uid)
    if not u:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    data = request.get_json()
    u.real_name = data.get('real_name', u.real_name)
    u.role = data.get('role', u.role)
    u.is_active = data.get('is_active', u.is_active)
    if data.get('password'):
        u.set_password(data['password'])
    db.session.commit()
    return jsonify({'success': True, 'data': u.to_dict()})


@app.route('/api/users/<int:uid>', methods=['DELETE'])
@admin_required
def delete_user(uid):
    u = User.query.get(uid)
    if not u:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    if u.username == 'admin':
        return jsonify({'success': False, 'message': '不能删除默认管理员'}), 400
    db.session.delete(u)
    db.session.commit()
    return jsonify({'success': True})


# ==================== API: 工序 ====================

@app.route('/api/processes', methods=['GET'])
@login_required
def get_processes():
    items = Process.query.all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/processes', methods=['POST'])
@admin_required
def create_process():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '工序名称不能为空'}), 400
    if Process.query.filter_by(name=name).first():
        return jsonify({'success': False, 'message': '工序名称已存在'}), 400
    p = Process(name=name, description=data.get('description', ''))
    db.session.add(p)
    db.session.commit()
    return jsonify({'success': True, 'data': p.to_dict()})


@app.route('/api/processes/<int:pid>', methods=['PUT'])
@admin_required
def update_process(pid):
    p = Process.query.get(pid)
    if not p:
        return jsonify({'success': False, 'message': '工序不存在'}), 404
    data = request.get_json()
    name = data.get('name', '').strip()
    if name and name != p.name:
        if Process.query.filter_by(name=name).first():
            return jsonify({'success': False, 'message': '工序名称已存在'}), 400
        p.name = name
    p.description = data.get('description', p.description)
    db.session.commit()
    return jsonify({'success': True, 'data': p.to_dict()})


@app.route('/api/processes/<int:pid>', methods=['DELETE'])
@admin_required
def delete_process(pid):
    p = Process.query.get(pid)
    if not p:
        return jsonify({'success': False, 'message': '工序不存在'}), 404
    if p.filter_models:
        return jsonify({'success': False, 'message': '该工序下存在物料型号，无法删除'}), 400
    db.session.delete(p)
    db.session.commit()
    return jsonify({'success': True})


# ==================== API: 物料型号 ====================

@app.route('/api/filter_models', methods=['GET'])
@login_required
def get_filter_models():
    q = FilterModel.query
    if request.args.get('system_only') in ('1', 'true'):
        q = q.filter_by(is_imported=False)
    items = q.all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/filter_models', methods=['POST'])
@admin_required
def create_filter_model():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '物料编码不能为空'}), 400
    raw_pd = data.get('production_date')
    pd_date = datetime.strptime(raw_pd, '%Y-%m-%d').date() if raw_pd else None
    fm = FilterModel(
        name=name,
        process_id=data.get('process_id'),
        max_sterilization_count=int(data['max_sterilization_count']) if data.get('max_sterilization_count') not in (None, '') else None,
        use_location=data.get('use_location', ''),
        production_date=pd_date
    )
    db.session.add(fm)
    db.session.commit()
    return jsonify({'success': True, 'data': fm.to_dict()})


@app.route('/api/filter_models/<int:fid>', methods=['PUT'])
@admin_required
def update_filter_model(fid):
    fm = FilterModel.query.get(fid)
    if not fm:
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    data = request.get_json()
    fm.name = data.get('name', fm.name).strip() or fm.name
    fm.process_id = data.get('process_id', fm.process_id)
    raw_s = data.get('max_sterilization_count')
    fm.max_sterilization_count = int(raw_s) if raw_s not in (None, '') else fm.max_sterilization_count
    fm.use_location = data.get('use_location', fm.use_location)
    raw_pd = data.get('production_date')
    if raw_pd:
        fm.production_date = datetime.strptime(raw_pd, '%Y-%m-%d').date()
    db.session.commit()
    return jsonify({'success': True, 'data': fm.to_dict()})


@app.route('/api/filter_models/<int:fid>', methods=['DELETE'])
@admin_required
def delete_filter_model(fid):
    fm = FilterModel.query.get(fid)
    if not fm:
        return jsonify({'success': False, 'message': '型号不存在'}), 404
    if fm.filters:
        return jsonify({'success': False, 'message': '该型号下存在物料实例，无法删除'}), 400
    db.session.delete(fm)
    db.session.commit()
    return jsonify({'success': True})


# ==================== API: 物料实例 ====================

@app.route('/api/filters', methods=['GET'])
@login_required
def get_filters():
    status = request.args.get('status', '')
    q = Filter.query
    if status:
        q = q.filter_by(status=status)
    items = q.order_by(Filter.created_at.desc()).all()
    result = []
    for f in items:
        f.update_status()
        result.append(f.to_dict())
    return jsonify(result)


@app.route('/api/filters/usage_records', methods=['GET'])
@login_required
def get_filter_usage_records():
    """返回仅通过'使用物料'手动添加的使用记录"""
    items = Filter.query.filter_by(is_usage_record=True).order_by(Filter.created_at.desc()).all()
    result = []
    for f in items:
        f.update_status()
        result.append(f.to_dict())
    return jsonify(result)


@app.route('/api/filters', methods=['POST'])
@admin_required
def create_filter():
    data = request.get_json()
    code = data.get('code', '').strip()
    if not code:
        return jsonify({'success': False, 'message': '物料编码不能为空'}), 400
    if Filter.query.filter_by(code=code).first():
        return jsonify({'success': False, 'message': '物料编码已存在'}), 400

    model_name = data.get('model_name', '').strip()
    if not model_name:
        return jsonify({'success': False, 'message': '型号不能为空'}), 400

    # 查找或创建 FilterModel
    fm = FilterModel.query.filter_by(name=model_name).first()
    if not fm:
        raw_s = data.get('max_sterilization_count')
        raw_d = data.get('max_days')
        fm = FilterModel(
            name=model_name,
            spec=data.get('spec', ''),
            max_sterilization_count=int(raw_s) if raw_s not in (None, '') else None,
            max_days=int(float(raw_d) * 365) if raw_d not in (None, '') else None,
            max_usage_count=0
        )
        db.session.add(fm)
        db.session.flush()
    else:
        fm.spec = data.get('spec', fm.spec)
        fm.max_sterilization_count = int(data.get('max_sterilization_count', fm.max_sterilization_count) or 0)
        fm.max_days = int(data.get('max_days', fm.max_days) or 0)

    f = Filter(
        code=code,
        model_id=fm.id,
        start_date=datetime.now().date(),
        is_usage_record=False
    )
    db.session.add(f)
    db.session.commit()
    return jsonify({'success': True, 'data': f.to_dict()})


@app.route('/api/filters/<int:fid>', methods=['GET'])
@login_required
def get_filter(fid):
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404
    f.update_status()
    return jsonify({'success': True, 'data': f.to_dict()})


@app.route('/api/filters/lookup')
@login_required
def lookup_filter():
    """根据物料编码查询物料信息（用于使用管理弹窗自动回填）"""
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'success': False, 'message': '请输入物料编码'}), 400
    f = Filter.query.filter_by(code=code).first()
    if not f:
        return jsonify({'success': False, 'message': '未找到该编码的物料'}), 404
    return jsonify({'success': True, 'data': f.to_dict()})


def _can_manage_filter(f):
    """行级权限：管理员、输入人、接收人均可管理该记录"""
    if not g.user:
        return False
    if g.user.role == 'admin':
        return True
    me = (g.user.real_name or g.user.username or '').strip()
    if f.operator and f.operator.strip() == me:
        return True
    if f.receivers:
        receivers = [r.strip() for r in f.receivers.split(',') if r.strip()]
        if me in receivers:
            return True
    return False


@app.route('/api/filters/<int:fid>', methods=['PUT'])
@login_required
def update_filter(fid):
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404
    if not _can_manage_filter(f):
        return jsonify({'success': False, 'message': '权限不足，仅管理员、输入人或接收人可编辑'}), 403
    data = request.get_json()
    is_admin = g.user and g.user.role == 'admin'

    # 管理员可以修改所有字段，操作员只能修改使用管理相关字段
    if is_admin:
        code = data.get('code', '').strip()
        if code and code != f.code:
            if Filter.query.filter_by(code=code).first():
                return jsonify({'success': False, 'message': '物料编码已存在'}), 400
            f.code = code

        model_name = data.get('model_name', '').strip()
        if model_name:
            fm = FilterModel.query.filter_by(name=model_name).first()
            if not fm:
                raw_s = data.get('max_sterilization_count')
                raw_d = data.get('max_days')
                fm = FilterModel(
                    name=model_name,
                    spec=data.get('spec', ''),
                    max_sterilization_count=int(raw_s) if raw_s not in (None, '') else None,
                    max_days=int(float(raw_d) * 365) if raw_d not in (None, '') else None,
                    max_usage_count=0
                )
                db.session.add(fm)
                db.session.flush()
            else:
                fm.spec = data.get('spec', fm.spec)
                raw_s = data.get('max_sterilization_count')
                fm.max_sterilization_count = int(raw_s) if raw_s not in (None, '') else fm.max_sterilization_count
                raw_d = data.get('max_days')
                fm.max_days = int(float(raw_d) * 365) if raw_d not in (None, '') else fm.max_days
            f.model_id = fm.id

    # 使用管理字段（管理员和操作员均可修改）
    if 'use_location' in data:
        f.use_location = data.get('use_location', f.use_location)
    if 'production_date' in data:
        raw_pd = data.get('production_date')
        f.production_date = datetime.strptime(raw_pd, '%Y-%m-%d').date() if raw_pd else f.production_date
    if 'operator' in data:
        f.operator = data.get('operator', f.operator)
    if 'receivers' in data:
        f.receivers = data.get('receivers', f.receivers)
    if 'use_process_name' in data:
        f.use_process_name = data.get('use_process_name', f.use_process_name)
    if 'record_max_sterilization' in data:
        raw = data.get('record_max_sterilization')
        f.record_max_sterilization = int(raw) if raw not in (None, '') else f.record_max_sterilization
    if 'current_sterilization_count' in data:
        raw = data.get('current_sterilization_count')
        f.current_sterilization_count = int(raw) if raw not in (None, '') else f.current_sterilization_count
    if 'is_usage_record' in data:
        f.is_usage_record = bool(data.get('is_usage_record', f.is_usage_record))

    db.session.commit()
    f.update_status()

    # 编辑后检查预警状态，如已恢复正常则重置推送记录
    alert_status = _get_usage_alert_status(f)
    if not alert_status['messages']:
        if f.last_sterilization_alert_at or f.last_expiry_alert_at:
            f.last_sterilization_alert_at = None
            f.last_expiry_alert_at = None
            db.session.commit()

    return jsonify({'success': True, 'data': f.to_dict()})


@app.route('/api/filters/<int:fid>', methods=['DELETE'])
@login_required
def delete_filter(fid):
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404
    if not _can_manage_filter(f):
        return jsonify({'success': False, 'message': '权限不足，仅管理员、输入人或接收人可删除'}), 403
    db.session.delete(f)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/filters/<int:fid>/use', methods=['POST'])
@login_required
def use_filter(fid):
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404
    data = request.get_json(silent=True) or {}
    f.current_usage_count += 1
    db.session.add(UsageRecord(filter_id=f.id, note=data.get('note', '')))
    db.session.commit()
    f.update_status()
    return jsonify({'success': True, 'data': f.to_dict()})


@app.route('/api/filters/<int:fid>/sterilize', methods=['POST'])
@login_required
def sterilize_filter(fid):
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404
    data = request.get_json(silent=True) or {}
    f.current_sterilization_count += 1
    db.session.add(SterilizationRecord(filter_id=f.id, note=data.get('note', '')))
    db.session.commit()
    f.update_status()
    return jsonify({'success': True, 'data': f.to_dict()})


@app.route('/api/filters/<int:fid>/reset', methods=['POST'])
@login_required
def reset_filter(fid):
    """更换物料：重置计数并更新启用日期，同时记录更换历史"""
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404

    # 记录更换历史
    record = FilterReplacementRecord(
        filter_id=f.id,
        old_start_date=f.start_date,
        new_start_date=datetime.now().date(),
        old_usage_count=f.current_usage_count,
        old_sterilization_count=f.current_sterilization_count,
        replaced_by=g.user.real_name or g.user.username if g.user else ''
    )
    db.session.add(record)

    f.current_usage_count = 0
    f.current_sterilization_count = 0
    f.start_date = datetime.now().date()
    db.session.commit()
    f.update_status()
    return jsonify({'success': True, 'data': f.to_dict()})


@app.route('/api/filters/<int:fid>/warnings', methods=['GET'])
@login_required
def get_filter_warnings(fid):
    f = Filter.query.get(fid)
    if not f:
        return jsonify({'success': False, 'message': '物料不存在'}), 404
    return jsonify({'success': True, 'warnings': f.get_warnings()})


# ==================== API: 记录查询 ====================

@app.route('/api/records/usage', methods=['GET'])
@login_required
def get_usage_records():
    items = UsageRecord.query.order_by(UsageRecord.usage_date.desc()).limit(500).all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/records/sterilization', methods=['GET'])
@login_required
def get_sterilization_records():
    items = SterilizationRecord.query.order_by(SterilizationRecord.sterilization_date.desc()).limit(500).all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/records/replacement', methods=['GET'])
@login_required
def get_replacement_records():
    items = FilterReplacementRecord.query.order_by(FilterReplacementRecord.replaced_at.desc()).limit(500).all()
    return jsonify([i.to_dict() for i in items])


# ==================== API: 设置 ====================

@app.route('/api/settings', methods=['GET'])
@login_required
def get_settings():
    cfg = Config.query.first()
    return jsonify(cfg.to_dict() if cfg else {})


@app.route('/api/settings', methods=['POST'])
@admin_required
def save_settings():
    data = request.get_json()
    cfg = Config.query.first()
    if not cfg:
        cfg = Config()
        db.session.add(cfg)
    cfg.feishu_webhook = data.get('feishu_webhook', '').strip()
    cfg.feishu_app_id = data.get('feishu_app_id', '').strip()
    cfg.feishu_app_secret = data.get('feishu_app_secret', '').strip()
    cfg.alert_days = int(data.get('alert_days', 7) or 7)
    cfg.alert_sterilization_remaining = int(data.get('alert_sterilization_remaining', 0) or 0)
    cfg.alert_push_enabled = bool(data.get('alert_push_enabled', False))
    cfg.alert_push_time = data.get('alert_push_time', '08:00').strip() or '08:00'
    cfg.alert_push_receivers = data.get('alert_push_receivers', '').strip()
    # 飞书多维表格同步配置
    cfg.feishu_bitable_app_token = data.get('feishu_bitable_app_token', '').strip()
    cfg.feishu_bitable_table_id = data.get('feishu_bitable_table_id', '').strip()
    cfg.feishu_bitable_sync_enabled = bool(data.get('feishu_bitable_sync_enabled', False))
    raw_interval = data.get('feishu_bitable_sync_interval', 5)
    cfg.feishu_bitable_sync_interval = int(raw_interval) if raw_interval else 5
    db.session.commit()
    # 重新配置定时同步任务
    _reschedule_bitable_sync()
    return jsonify({'success': True})


# ==================== API: 飞书多维表格同步 ====================

_bitable_sync_job_id = 'feishu_bitable_sync'


def _reschedule_bitable_sync():
    """根据配置重新调度飞书多维表格同步任务"""
    # Serverless 环境不支持常驻后台任务
    if is_vercel:
        return

    cfg = Config.query.first()
    interval = cfg.feishu_bitable_sync_interval if cfg else 5
    enabled = cfg.feishu_bitable_sync_enabled if cfg else False

    if scheduler.get_job(_bitable_sync_job_id):
        scheduler.remove_job(_bitable_sync_job_id)

    if enabled:
        scheduler.add_job(
            auto_sync_feishu_bitable,
            'interval',
            minutes=max(1, interval),
            id=_bitable_sync_job_id,
            replace_existing=True
        )
        print(f'[Scheduler] 飞书多维表格同步任务已启动，间隔 {interval} 分钟')


@app.route('/api/feishu/bitable/sync', methods=['POST'])
@admin_required
def manual_sync_feishu_bitable():
    """手动触发飞书多维表格同步"""
    result = sync_from_feishu_bitable()
    return jsonify(result)


@app.route('/api/feishu/bitable/test', methods=['POST'])
@admin_required
def test_feishu_bitable_connection():
    """测试飞书多维表格连接"""
    cfg = Config.query.first()
    if not cfg or not cfg.feishu_app_id or not cfg.feishu_app_secret:
        return jsonify({'success': False, 'message': '未配置飞书应用 App ID 和 App Secret'}), 400
    if not cfg.feishu_bitable_app_token:
        return jsonify({'success': False, 'message': '未配置多维表格 App Token'}), 400

    tenant_token = _get_feishu_tenant_access_token(cfg.feishu_app_id, cfg.feishu_app_secret)
    if not tenant_token:
        return jsonify({'success': False, 'message': '无法获取 tenant_access_token，请检查 App ID 和 App Secret'}), 400

    tables = _get_feishu_bitable_tables(cfg.feishu_bitable_app_token, tenant_token)
    if not tables:
        return jsonify({'success': False, 'message': '无法获取表格列表，请检查 App Token 是否正确'}), 400

    table_id = cfg.feishu_bitable_table_id
    if not table_id:
        table_id = tables[0][0]
        cfg.feishu_bitable_table_id = table_id
        db.session.commit()

    # 尝试读取一条记录
    records = _get_feishu_bitable_records(cfg.feishu_bitable_app_token, table_id, tenant_token, page_size=1)
    sample = records[0].get('fields', {}) if records else {}

    return jsonify({
        'success': True,
        'message': '连接成功',
        'tables': [{'table_id': t[0], 'name': t[1]} for t in tables],
        'current_table_id': table_id,
        'sample_fields': list(sample.keys())
    })


# ==================== API: 预警与飞书推送 ====================

@app.route('/api/alert_summary', methods=['GET'])
@login_required
def alert_summary():
    # 预警看板只显示使用管理中的物料（is_usage_record=True）
    filters = Filter.query.filter_by(is_usage_record=True).all()
    expired = []
    warning = []
    normal = []
    for f in filters:
        f.update_status()
        d = f.to_dict()
        d['warnings'] = f.get_warnings()
        if f.status == 'expired':
            expired.append(d)
        elif f.status == 'warning':
            warning.append(d)
        else:
            normal.append(d)
    return jsonify({
        'expired': expired,
        'warning': warning,
        'normal': normal,
        'counts': {
            'total': len(filters),
            'expired': len(expired),
            'warning': len(warning),
            'normal': len(normal)
        }
    })


def build_feishu_card(expired, warning):
    elements = []
    if expired:
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": f"**🔴 需立即更换（{len(expired)}支）**"}
        })
        for f in expired:
            warn_text = '；'.join(f['warnings'])
            elements.append({
                "tag": "div",
                "text": {"tag": "lark_md", "content": f"• **{f['code']}** | {f['process_name']} | {f['model_name']} | {warn_text}"}
            })
    if warning:
        if expired:
            elements.append({"tag": "hr"})
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": f"**🟡 即将到期（{len(warning)}支）**"}
        })
        for f in warning:
            warn_text = '；'.join(f['warnings'])
            elements.append({
                "tag": "div",
                "text": {"tag": "lark_md", "content": f"• **{f['code']}** | {f['process_name']} | {f['model_name']} | {warn_text}"}
            })
    if not expired and not warning:
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": "✅ 当前所有物料均正常，无需更换。"}
        })

    card = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"物料更换提醒 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"},
            "template": "red" if expired else ("yellow" if warning else "green")
        },
        "elements": elements
    }
    return card


@app.route('/api/send_feishu', methods=['POST'])
@login_required
def send_feishu():
    cfg = Config.query.first()
    if not cfg or not cfg.feishu_webhook:
        return jsonify({'success': False, 'message': '请先配置飞书 Webhook'}), 400

    data = request.get_json(silent=True) or {}
    filters = Filter.query.all()
    expired = []
    warning = []
    for f in filters:
        f.update_status()
        d = f.to_dict()
        d['warnings'] = f.get_warnings()
        if f.status == 'expired':
            expired.append(d)
        elif f.status == 'warning':
            warning.append(d)

    only_alert = data.get('only_alert', False)
    if only_alert and not expired and not warning:
        return jsonify({'success': True, 'message': '当前没有异常，无需发送'})

    card = build_feishu_card(expired, warning)
    payload = {"msg_type": "interactive", "card": card}

    try:
        resp = requests.post(cfg.feishu_webhook, json=payload, timeout=15)
        resp_json = resp.json()
        if resp_json.get('code') == 0:
            return jsonify({'success': True, 'message': '飞书消息发送成功'})
        else:
            return jsonify({'success': False, 'message': f"飞书接口错误: {resp_json}"}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'请求异常: {str(e)}'}), 500


# ==================== API: 数据导入 ====================

@app.route('/api/import/filter_models', methods=['POST'])
@admin_required
def import_filter_models():
    """从上传的 Excel 文件导入物料数据，同时创建/更新物料型号和物料实例"""
    uploaded_file = request.files.get('file')
    if not uploaded_file:
        return jsonify({'success': False, 'message': '请先选择要导入的文件'}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'success': False, 'message': '缺少 openpyxl，无法读取 Excel'}), 500

    try:
        file_stream = io.BytesIO(uploaded_file.read())
        wb = load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'success': False, 'message': f'读取 Excel 失败: {str(e)}'}), 500

    # 读取表头并建立列映射
    headers = []
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] is not None:
            headers = [str(h).strip() if h else '' for h in row]
            header_row = idx
            break

    if not header_row:
        return jsonify({'success': False, 'message': '无法识别表头行'}), 400

    # 支持的列名映射（兼容不同版本的模板）
    col_map = {}
    for idx, h in enumerate(headers):
        if '物料名称' in h:
            col_map['name'] = idx
        elif '型号' in h and '规格' in h:
            col_map['spec'] = idx
        elif '有效期' in h:
            col_map['max_days'] = idx
        elif h == '物料代码':
            col_map['code'] = idx
        elif '供应商' in h and '名称' in h:
            col_map['supplier'] = idx
        elif '灭菌' in h and '次数' in h:
            col_map['max_sterilization_count'] = idx

    if 'name' not in col_map:
        return jsonify({'success': False, 'message': f'Excel 中缺少「物料名称」列'}), 400

    fm_created = 0
    fm_updated = 0
    f_created = 0
    f_updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = str(row[col_map['name']]).strip() if col_map['name'] < len(row) and row[col_map['name']] is not None else ''
        if not name:
            skipped += 1
            continue

        # 物料编码（物料代码）
        code = str(row[col_map['code']]).strip() if 'code' in col_map and col_map['code'] < len(row) and row[col_map['code']] is not None else ''
        if not code:
            skipped += 1
            continue

        spec = str(row[col_map['spec']]).strip() if 'spec' in col_map and col_map['spec'] < len(row) and row[col_map['spec']] is not None else ''
        supplier = str(row[col_map['supplier']]).strip() if 'supplier' in col_map and col_map['supplier'] < len(row) and row[col_map['supplier']] is not None else ''

        # 规格拼接供应商信息
        full_spec = spec
        if supplier:
            full_spec = f"{spec} | 供应商:{supplier}" if spec else f"供应商:{supplier}"

        # 有效期（空值保留为 NULL）
        max_days = None
        if 'max_days' in col_map and col_map['max_days'] < len(row) and row[col_map['max_days']] is not None:
            try:
                raw_val = float(row[col_map['max_days']])
                if raw_val > 0:
                    max_days = int(raw_val * 365)
            except (ValueError, TypeError):
                max_days = None

        # 可灭菌次数（空值保留为 NULL）
        max_sterilization_count = None
        if 'max_sterilization_count' in col_map and col_map['max_sterilization_count'] < len(row) and row[col_map['max_sterilization_count']] is not None:
            raw = str(row[col_map['max_sterilization_count']]).strip()
            if raw:
                try:
                    max_sterilization_count = int(float(raw))
                except (ValueError, TypeError):
                    max_sterilization_count = None

        # 1. 创建/更新物料型号（标记为导入）
        fm = FilterModel.query.filter_by(name=name).first()
        if fm:
            fm.spec = full_spec
            fm.max_days = max_days if max_days is not None else fm.max_days
            fm.max_sterilization_count = max_sterilization_count if max_sterilization_count is not None else fm.max_sterilization_count
            fm.is_imported = True
            fm_updated += 1
        else:
            fm = FilterModel(
                name=name,
                spec=full_spec,
                max_days=max_days,
                max_usage_count=0,
                max_sterilization_count=max_sterilization_count,
                is_imported=True
            )
            db.session.add(fm)
            db.session.flush()  # 获取 fm.id
            fm_created += 1

        # 2. 创建/更新物料实例
        f = Filter.query.filter_by(code=code).first()
        if f:
            f.model_id = fm.id
            f_updated += 1
        else:
            f = Filter(
                code=code,
                model_id=fm.id,
                start_date=datetime.now().date(),
                is_usage_record=False
            )
            db.session.add(f)
            f_created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'数据库保存失败: {str(e)}'}), 500

    msg = f'导入完成：型号新增 {fm_created} 条/更新 {fm_updated} 条，物料新增 {f_created} 条/更新 {f_updated} 条，跳过 {skipped} 条'
    return jsonify({
        'success': True,
        'message': msg,
        'data': {
            'fm_created': fm_created, 'fm_updated': fm_updated,
            'f_created': f_created, 'f_updated': f_updated,
            'skipped': skipped
        }
    })


# ==================== API: 数据导出 ====================

def create_excel_response(wb, filename):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/filters')
@login_required
def export_filters():
    wb = Workbook()
    ws = wb.active
    ws.title = '数据库'
    # 与导入文件格式保持一致
    headers = ['物料代码', '物料名称', '型号/规格', '单位', '可灭菌次数', '有效期(年)', '供应商名称']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

    items = Filter.query.all()
    for f in items:
        fm = f.model
        if not fm:
            continue

        # 从 spec 中拆分规格和供应商
        full_spec = fm.spec or ''
        spec_part = full_spec
        supplier_name = ''
        if ' | 供应商:' in full_spec:
            spec_part, supplier_part = full_spec.split(' | 供应商:', 1)
            supplier_name = supplier_part
        elif full_spec.startswith('供应商:'):
            spec_part = ''
            supplier_name = full_spec[4:]

        # 有效期从天还原为年（与导入文件一致）
        expire_years = ''
        if fm.max_days and fm.max_days > 0:
            expire_years = round(fm.max_days / 365, 1)
            # 如果是整数则显示整数
            if expire_years == int(expire_years):
                expire_years = int(expire_years)

        ws.append([
            f.code,
            fm.name,
            spec_part,
            '支',
            fm.max_sterilization_count if fm.max_sterilization_count is not None else '',
            expire_years,
            supplier_name
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    return create_excel_response(wb, f'数据库_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/api/export/alerts')
@login_required
def export_alerts():
    wb = Workbook()
    # 需更换
    ws1 = wb.active
    ws1.title = '需更换'
    ws1.append(['物料编码', '工序', '型号', '规格', '预警原因', '启用日期'])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

    # 即将到期
    ws2 = wb.create_sheet('即将到期')
    ws2.append(['物料编码', '工序', '型号', '规格', '预警原因', '启用日期'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

    filters = Filter.query.all()
    for f in filters:
        f.update_status()
        d = f.to_dict()
        d['warnings'] = f.get_warnings()
        if f.status == 'expired':
            ws1.append([d['code'], d['process_name'], d['model_name'], d['spec'] or '',
                        '；'.join(d['warnings']), d['start_date']])
        elif f.status == 'warning':
            ws2.append([d['code'], d['process_name'], d['model_name'], d['spec'] or '',
                        '；'.join(d['warnings']), d['start_date']])

    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    return create_excel_response(wb, f'预警清单_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/api/export/records')
@login_required
def export_records():
    record_type = request.args.get('type', 'usage')
    wb = Workbook()
    ws = wb.active

    if record_type == 'usage':
        ws.title = '使用记录'
        ws.append(['物料编码', '工序', '型号', '登记时间', '备注'])
        items = UsageRecord.query.order_by(UsageRecord.usage_date.desc()).all()
        for i in items:
            d = i.to_dict()
            ws.append([d['filter_code'], d['process_name'], d['filter_model'], d['usage_date'], d['note']])
    elif record_type == 'sterilization':
        ws.title = '灭菌记录'
        ws.append(['物料编码', '工序', '型号', '登记时间', '备注'])
        items = SterilizationRecord.query.order_by(SterilizationRecord.sterilization_date.desc()).all()
        for i in items:
            d = i.to_dict()
            ws.append([d['filter_code'], d['process_name'], d['filter_model'], d['sterilization_date'], d['note']])
    elif record_type == 'replacement':
        ws.title = '更换记录'
        ws.append(['物料编码', '工序', '型号', '旧启用日期', '新启用日期', '旧使用次数', '旧灭菌次数', '更换时间', '操作人', '备注'])
        items = FilterReplacementRecord.query.order_by(FilterReplacementRecord.replaced_at.desc()).all()
        for i in items:
            d = i.to_dict()
            ws.append([d['filter_code'], d['process_name'], d['filter_model'],
                       d['old_start_date'], d['new_start_date'],
                       d['old_usage_count'], d['old_sterilization_count'],
                       d['replaced_at'], d['replaced_by'], d['note']])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    return create_excel_response(wb, f'{ws.title}_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ==================== API: 报表统计 ====================

@app.route('/api/reports/summary')
@login_required
def report_summary():
    # 总体统计
    total = Filter.query.count()
    expired = Filter.query.filter_by(status='expired').count()
    warning = Filter.query.filter_by(status='warning').count()
    normal = Filter.query.filter_by(status='normal').count()

    # 按工序统计
    process_stats = []
    for p in Process.query.all():
        filter_count = 0
        p_expired = 0
        p_warning = 0
        for fm in p.filter_models:
            for f in fm.filters:
                filter_count += 1
                f.update_status()
                if f.status == 'expired':
                    p_expired += 1
                elif f.status == 'warning':
                    p_warning += 1
        process_stats.append({
            'process_name': p.name,
            'total': filter_count,
            'expired': p_expired,
            'warning': p_warning,
            'normal': filter_count - p_expired - p_warning
        })

    # 按型号统计
    model_stats = []
    for m in FilterModel.query.all():
        count = len(m.filters)
        if count > 0:
            model_stats.append({
                'model_name': m.name,
                'process_name': m.process.name if m.process else '',
                'total': count
            })

    # 最近30天使用/灭菌/更换次数
    thirty_days_ago = datetime.now() - timedelta(days=30)
    usage_30d = UsageRecord.query.filter(UsageRecord.usage_date >= thirty_days_ago).count()
    sterilization_30d = SterilizationRecord.query.filter(SterilizationRecord.sterilization_date >= thirty_days_ago).count()
    replacement_30d = FilterReplacementRecord.query.filter(FilterReplacementRecord.replaced_at >= thirty_days_ago).count()

    # 更换历史平均使用次数和灭菌次数
    replacements = FilterReplacementRecord.query.all()
    avg_usage = 0
    avg_sterilization = 0
    if replacements:
        avg_usage = round(sum(r.old_usage_count for r in replacements) / len(replacements), 1)
        avg_sterilization = round(sum(r.old_sterilization_count for r in replacements) / len(replacements), 1)

    return jsonify({
        'overview': {'total': total, 'expired': expired, 'warning': warning, 'normal': normal},
        'process_stats': process_stats,
        'model_stats': model_stats,
        'activity_30d': {'usage': usage_30d, 'sterilization': sterilization_30d, 'replacement': replacement_30d},
        'replacement_avg': {'avg_usage_count': avg_usage, 'avg_sterilization_count': avg_sterilization}
    })


@app.route('/api/reports/monthly')
@login_required
def report_monthly():
    """最近12个月的使用/灭菌/更换趋势"""
    now = datetime.now()
    months = []
    for i in range(11, -1, -1):
        d = now - timedelta(days=30*i)
        start = d.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i > 0:
            end = (start + timedelta(days=32)).replace(day=1)
        else:
            end = now
        label = start.strftime('%Y-%m')
        usage = UsageRecord.query.filter(UsageRecord.usage_date >= start, UsageRecord.usage_date < end).count()
        sterilization = SterilizationRecord.query.filter(SterilizationRecord.sterilization_date >= start, SterilizationRecord.sterilization_date < end).count()
        replacement = FilterReplacementRecord.query.filter(FilterReplacementRecord.replaced_at >= start, FilterReplacementRecord.replaced_at < end).count()
        months.append({'month': label, 'usage': usage, 'sterilization': sterilization, 'replacement': replacement})
    return jsonify(months)


# ==================== 入口 ====================

if __name__ == '__main__':
    with app.app_context():
        _reschedule_bitable_sync()
    app.run(host='0.0.0.0', port=5000, debug=True)
